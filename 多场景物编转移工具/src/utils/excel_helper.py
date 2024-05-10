# -*- coding: utf-8 -*-
"""
@file：csv_helper.py
@description: excel文件处理
@author: longyunyue
@date：2021/9/3
"""
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font


class WorkBookWrapper(object):

	def __init__(self, name='workbook', wb=None):
		self._file_name = name
		if not wb:
			self._wb = Workbook()
		else:
			self._wb = wb

	@property
	def work_book(self):
		return self._wb

	@property
	def active_sheet(self):
		return self._wb.active

	def set_active_sheet(self, sheet_or_index):
		self._wb.active = sheet_or_index

	def get_sheet(self, name):
		return self._wb[name]

	def create_sheet(self, name=None, index=None):
		return self._wb.create_sheet(name, index)

	@property
	def file_name(self):
		return self._file_name

	@file_name.setter
	def file_name(self, new):
		self._file_name = new

	def save(self, file_path=None):
		self._wb.save(file_path)

	def _adjust_column_width(self):
		for sheet in self.work_book.worksheets:
			for col in sheet.columns:
				max_length = 0
				column = col[0].column
				for cell in col:
					try:
						if len(str(cell.value)) > max_length:
							max_length = len(cell.value)
					except:
						pass
				adjusted_width = (max_length + 2) * 1.2
				sheet.column_dimensions[column].width = adjusted_width

	def write_dict_rows_to_sheet(self, sheet_name, dict_rows):
		"""
		写入字典数据，第一个元素的values()是描述，后面的是值
		"""
		if sheet_name not in self._wb:
			sheet = self.create_sheet(sheet_name)
		else:
			sheet = self._wb[sheet_name]
		self.set_active_sheet(sheet)
		col = 1
		key_to_col = {}
		key_desc_dict = dict_rows[0]
		# 写入第一行（key）,第二行（desc）
		for key, desc in key_desc_dict.iteritems():
			key_cell = sheet.cell(row=1, column=col, value=key)
			key_cell.fill = PatternFill("solid", fgColor="DDDD00")
			key_cell.font = Font(b=True, color="000000")
			desc_cell = sheet.cell(row=2, column=col, value=desc)
			desc_cell.fill = PatternFill("solid", fgColor="00DD00")
			desc_cell.font = Font(b=True, color="000000")
			key_to_col[key] = col
			col += 1
		# 写入剩下的数据行
		row = 3
		for d in dict_rows[1:]:
			for key, value in d.iteritems():
				value_col = key_to_col[key]
				data_cell = sheet.cell(row=row, column=value_col, value=str(value))
				data_cell.fill = PatternFill("solid", fgColor="DDDDDD")
			row += 1
		# 调整列宽
		self._adjust_column_width()

	def write_list_rows_to_sheet(self, sheet_name, list_data):
		"""写入列表数据"""
		if sheet_name not in self._wb:
			sheet = self.create_sheet(sheet_name)
		else:
			sheet = self._wb[sheet_name]
		self.set_active_sheet(sheet)
		for row in xrange(len(list_data)):
			for col in xrange(len(list_data[row])):
				cell = sheet.cell(row=row + 1, column=col + 1, value=list_data[row][col])
				# cell.fill = PatternFill("solid", fgColor="DDDD00")
				# cell.font = Font(b=True, color="000000")
		# self._adjust_column_width()

	def read_csv_dict(self):
		"""
		数据读成字典
		:return:
		"""
		result = {}
		for name in self.work_book.sheetnames:
			sheet_data = []
			col_to_key = {}
			sheet = self.get_sheet(name)
			for row in range(sheet.min_row, sheet.max_row + 1):
				dict_row = {}
				for col in range(sheet.min_column, sheet.max_column + 1):
					cell = sheet.cell(row=row, column=col)
					v = cell.value
					if v is None:
						v = ''
					if isinstance(v, unicode):
						v = v.encode('utf-8')
					if row == 1:
						col_to_key[col] = v
					else:
						dict_row[col_to_key[col]] = v
				dict_row and sheet_data.append(dict_row)
			result[name] = sheet_data
		return result

	def remove_default_ws(self):
		ins = self._wb
		ws = ins['Sheet']
		ins.remove(ws)


def write_csv_dict_to_excel(file_name, sheet_info, post_process=None, dir_path=None):
	"""
	写入csv的dict_rows到excel文件
	:param file_name: 文件名称
	:param sheet_info: {sheet_name: dict_rows}
	:return:
	"""
	wb = WorkBookWrapper(file_name)
	# 导出后处理
	if post_process:
		sheet_info = post_process(sheet_info)
	for sheet_name, data in sheet_info.iteritems():
		wb.write_dict_rows_to_sheet(sheet_name, data)
	wb.remove_default_ws()
	path = None if dir_path is None else os.path.join(dir_path, file_name)
	wb.save(path)


def write_list_to_excel(file_name, sheet_info, post_process=None, dir_path=None):
	"""将list写入到文件"""
	file_name = u"%s" % file_name
	wb = WorkBookWrapper(file_name)
	# 导出后处理
	if post_process:
		sheet_info = post_process(sheet_info)
	for sheet_name, data in sheet_info.iteritems():
		sheet_name = u"%s" % sheet_name
		wb.write_list_rows_to_sheet(sheet_name, data)
	wb.remove_default_ws()
	path = None if dir_path is None else os.path.join(dir_path, file_name)
	wb.save(path)


def read_csv_dict_from_excel(file_path):
	"""
	从excel读取csv格式的数据
	:param file_path: 文件路径
	:param sheet_names: 表名数组
	:return:
	"""
	wb = load_workbook(file_path)
	if wb:
		wb_wrapper = WorkBookWrapper(wb=wb)
		return wb_wrapper.read_csv_dict()






